import sys
sys.stdout.reconfigure(encoding='utf-8')

from pathlib import Path
import warnings

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

warnings.filterwarnings('ignore')

plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

DATA_DIR = Path('附件数据')
FIGURE_DIR = Path('output/figures')
FIGURE_DIR.mkdir(parents=True, exist_ok=True)

FILES = {
    'GPU_information': DATA_DIR / 'GPU_information.xlsx',
    'network_latency': DATA_DIR / 'network_latency.xlsx',
    'power_mapping': DATA_DIR / 'power_mapping.xlsx',
    'region_time_data': DATA_DIR / 'region_time_data.xlsx',
    'storage_information': DATA_DIR / 'storage_information.xlsx',
    'workload_trace': DATA_DIR / 'workload_trace.xlsx',
}

IDENTIFIER_KEYWORDS = (
    'id', '编号', 'region', '区域', 'group', '组别',
    'material', '材料', 'catalyst', '催化剂'
)


def print_section(title):
    print('\n' + '=' * 80)
    print(title)
    print('=' * 80)


def detect_merged_identifier_columns(file_path, sheet_name, columns):
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        worksheet = workbook[sheet_name]
        merged_columns = set()

        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row <= 1 <= merged_range.max_row:
                continue
            for column_number in range(merged_range.min_col, merged_range.max_col + 1):
                header = worksheet.cell(row=1, column=column_number).value
                if header in columns:
                    header_text = str(header).lower()
                    if any(keyword in header_text for keyword in IDENTIFIER_KEYWORDS):
                        merged_columns.add(header)

        workbook.close()
        return sorted(merged_columns)
    except Exception as exc:
        print(f'合并单元格检测失败: 文件={file_path}, 表单={sheet_name}, 原因={exc}')
        return []


def load_excel_workbook(name, file_path):
    print_section(f'读取文件: {file_path}')
    excel_file = pd.ExcelFile(file_path)
    print(f'表单数量: {len(excel_file.sheet_names)}')
    print(f'表单名称: {excel_file.sheet_names}')

    workbook_data = {}
    for sheet_name in excel_file.sheet_names:
        frame = pd.read_excel(file_path, sheet_name=sheet_name)
        merged_id_columns = detect_merged_identifier_columns(
            file_path, sheet_name, frame.columns
        )
        if merged_id_columns:
            print(
                f'表单 {sheet_name} 检测到标识列合并单元格,'
                f'执行前向填充: {merged_id_columns}'
            )
            frame[merged_id_columns] = frame[merged_id_columns].ffill()

        workbook_data[sheet_name] = frame
        print_table_overview(name, sheet_name, frame)

    return workbook_data


def print_table_overview(file_name, sheet_name, frame):
    print_section(f'数据概览: {file_name} / {sheet_name}')
    print(f'行数: {frame.shape[0]}')
    print(f'列数: {frame.shape[1]}')
    print(f'列名: {frame.columns.tolist()}')
    print('数据类型:')
    print(frame.dtypes.to_string())

    print('缺失值统计:')
    missing = frame.isna().sum()
    print(missing.to_string())
    print(f'缺失值总数: {int(missing.sum())}')
    print(f'重复行数: {int(frame.duplicated().sum())}')

    numeric = frame.select_dtypes(include=np.number)
    if numeric.empty:
        print('数值列描述统计: 无数值列')
    else:
        print('数值列描述统计:')
        print(
            numeric.describe(
                percentiles=[0.01, 0.05, 0.25, 0.5, 0.75, 0.95, 0.99]
            ).T.to_string()
        )

    categorical = frame.select_dtypes(
        include=['object', 'category', 'string', 'bool']
    )
    if categorical.empty:
        print('类别列分布: 无类别列')
    else:
        print('类别列分布:')
        for column in categorical.columns:
            counts = frame[column].astype('string').fillna('<缺失>').value_counts(
                dropna=False
            )
            print(f'列 {column}: 唯一值数量={frame[column].nunique(dropna=True)}')
            print(counts.head(30).to_string())


def save_figure(fig, filename):
    path = FIGURE_DIR / filename
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f'图表已保存: {path}')


def print_iqr_outliers(frame, columns, group_column=None, label=''):
    print_section(f'IQR异常值检测: {label}')
    for column in columns:
        if column not in frame.columns:
            continue

        if group_column and group_column in frame.columns:
            for group_value, group in frame.groupby(group_column, dropna=False):
                values = pd.to_numeric(group[column], errors='coerce').dropna()
                if values.empty:
                    continue
                q1 = values.quantile(0.25)
                q3 = values.quantile(0.75)
                iqr = q3 - q1
                lower = q1 - 1.5 * iqr
                upper = q3 + 1.5 * iqr
                count = int(((values < lower) | (values > upper)).sum())
                print(
                    f'列={column}, {group_column}={group_value}, '
                    f'Q1={q1}, Q3={q3}, IQR={iqr}, '
                    f'下界={lower}, 上界={upper}, 异常值数量={count}'
                )
        else:
            values = pd.to_numeric(frame[column], errors='coerce').dropna()
            if values.empty:
                continue
            q1 = values.quantile(0.25)
            q3 = values.quantile(0.75)
            iqr = q3 - q1
            lower = q1 - 1.5 * iqr
            upper = q3 + 1.5 * iqr
            count = int(((values < lower) | (values > upper)).sum())
            print(
                f'列={column}, Q1={q1}, Q3={q3}, IQR={iqr}, '
                f'下界={lower}, 上界={upper}, 异常值数量={count}'
            )


def plot_numeric_histograms(frame, columns, filename, title):
    valid_columns = [column for column in columns if column in frame.columns]
    if not valid_columns:
        return

    ncols = 3
    nrows = int(np.ceil(len(valid_columns) / ncols))
    fig, axes = plt.subplots(
        nrows, ncols, figsize=(6 * ncols, 4 * nrows), squeeze=False
    )
    for axis, column in zip(axes.flat, valid_columns):
        values = pd.to_numeric(frame[column], errors='coerce').dropna()
        axis.hist(values, bins=30, color='#4472C4', alpha=0.8, edgecolor='white')
        axis.set_title(column)
        axis.set_xlabel('数值')
        axis.set_ylabel('频数')

    for axis in axes.flat[len(valid_columns):]:
        axis.axis('off')

    fig.suptitle(title, fontsize=16)
    save_figure(fig, filename)


def plot_correlation(frame, columns, filename, title):
    valid_columns = [
        column for column in columns
        if column in frame.columns and pd.api.types.is_numeric_dtype(frame[column])
    ]
    if len(valid_columns) < 2:
        return

    correlation = frame[valid_columns].corr()
    print_section(f'相关性矩阵: {title}')
    print(correlation.to_string())

    fig, axis = plt.subplots(figsize=(12, 10))
    image = axis.imshow(correlation, cmap='coolwarm', vmin=-1, vmax=1)
    axis.set_xticks(range(len(valid_columns)))
    axis.set_yticks(range(len(valid_columns)))
    axis.set_xticklabels(valid_columns, rotation=90)
    axis.set_yticklabels(valid_columns)
    fig.colorbar(image, ax=axis, label='相关系数')
    axis.set_title(title)
    save_figure(fig, filename)


def analyze_gpu_information(frame):
    print_section('GPU中心容量与功率约束分析')
    selected = [
        'Region', 'Total_GPU', 'Available_GPU', 'Reserved_GPU_Ratio',
        'Max_IT_Power_MW', 'Max_Facility_Power_MW', 'PUE',
        'Max_Workload_GPUh_per_h'
    ]
    print(frame[selected].to_string(index=False))

    calculated_available = frame['Total_GPU'] * (1 - frame['Reserved_GPU_Ratio'])
    difference = frame['Available_GPU'] - calculated_available
    print('Available_GPU规则复核:')
    for region, actual, expected, diff in zip(
        frame['Region'], frame['Available_GPU'], calculated_available, difference
    ):
        print(
            f'Region={region}, 文件值={actual}, 计算值={expected}, 差值={diff}'
        )

    fig, axis = plt.subplots(figsize=(10, 6))
    x = np.arange(len(frame))
    width = 0.35
    axis.bar(x - width / 2, frame['Total_GPU'], width, label='Total_GPU')
    axis.bar(x + width / 2, frame['Available_GPU'], width, label='Available_GPU')
    axis.set_xticks(x)
    axis.set_xticklabels(frame['Region'])
    axis.set_ylabel('GPU等效单元')
    axis.set_title('各区域GPU容量')
    axis.legend()
    save_figure(fig, 'eda_gpu_capacity_by_region.png')

    print_iqr_outliers(
        frame,
        [
            'Total_GPU', 'Available_GPU', 'Max_IT_Power_MW',
            'Max_Facility_Power_MW', 'PUE'
        ],
        label='GPU中心基础情况'
    )


def analyze_network_latency(frame):
    print_section('网络时延分析')
    directional_pairs = frame.merge(
        frame,
        left_on=['FromRegion', 'ToRegion'],
        right_on=['ToRegion', 'FromRegion'],
        suffixes=('_forward', '_reverse')
    )
    directional_pairs['LatencyDifference_ms'] = (
        directional_pairs['NetworkLatency_ms_forward']
        - directional_pairs['NetworkLatency_ms_reverse']
    )
    print('双向时延差异统计:')
    print(directional_pairs['LatencyDifference_ms'].describe().to_string())
    print(
        f'非对称有向边数量: '
        f'{int((directional_pairs["LatencyDifference_ms"] != 0).sum())}'
    )

    latency_summary = frame.groupby('FromRegion')['NetworkLatency_ms'].agg(
        ['count', 'min', 'max', 'mean', 'median', 'std']
    )
    print('按来源区域的网络时延统计:')
    print(latency_summary.to_string())

    class_distribution = frame['LatencyClass'].value_counts(dropna=False)
    print('时延类别分布:')
    print(class_distribution.to_string())

    matrix = frame.pivot(
        index='FromRegion',
        columns='ToRegion',
        values='NetworkLatency_ms'
    )
    fig, axis = plt.subplots(figsize=(8, 7))
    image = axis.imshow(matrix.values, cmap='YlOrRd')
    axis.set_xticks(range(len(matrix.columns)))
    axis.set_yticks(range(len(matrix.index)))
    axis.set_xticklabels(matrix.columns)
    axis.set_yticklabels(matrix.index)
    axis.set_xlabel('执行区域')
    axis.set_ylabel('来源区域')
    axis.set_title('区域间单向网络时延矩阵')
    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            axis.text(j, i, f'{matrix.iloc[i, j]:.0f}', ha='center', va='center')
    fig.colorbar(image, ax=axis, label='毫秒')
    save_figure(fig, 'eda_network_latency_heatmap.png')

    print_iqr_outliers(
        frame, ['NetworkLatency_ms'], group_column='FromRegion',
        label='按来源区域的网络时延'
    )


def analyze_power_mapping(frame):
    print_section('任务功率映射分析')
    print(frame.to_string(index=False))

    fig, axis = plt.subplots(figsize=(9, 5))
    axis.bar(
        frame['TaskType'],
        frame['GPU_Power_MW_per_EquivalentGPU'],
        color=['#C0504D', '#F2A900', '#70AD47']
    )
    axis.set_ylabel('MW/等效GPU')
    axis.set_title('不同任务类型的单位等效GPU功率')
    axis.tick_params(axis='x', rotation=15)
    save_figure(fig, 'eda_task_power_mapping.png')


def analyze_storage_information(frame):
    print_section('储能参数分析')
    selected = [
        'Region', 'StorageCapacity_MWh', 'MinSOC_MWh', 'InitialSOC_MWh',
        'MaxChargePower_MW', 'MaxDischargePower_MW',
        'ChargeEfficiency', 'DischargeEfficiency',
        'MaxGridImport_MW', 'MaxGridExport_MW'
    ]
    print(frame[selected].to_string(index=False))

    checks = pd.DataFrame({
        'Region': frame['Region'],
        'InitialSOC_not_below_min':
            frame['InitialSOC_MWh'] >= frame['MinSOC_MWh'],
        'InitialSOC_not_above_capacity':
            frame['InitialSOC_MWh'] <= frame['StorageCapacity_MWh'],
        'SellLimit_matches_export':
            frame['SellLimit_MW'] == frame['MaxGridExport_MW'],
        'ChargePower_nonnegative':
            frame['MaxChargePower_MW'] >= 0,
        'DischargePower_nonnegative':
            frame['MaxDischargePower_MW'] >= 0,
    })
    print('储能边界规则复核:')
    print(checks.to_string(index=False))

    fig, axis = plt.subplots(figsize=(11, 6))
    x = np.arange(len(frame))
    width = 0.25
    axis.bar(
        x - width, frame['StorageCapacity_MWh'], width,
        label='StorageCapacity_MWh'
    )
    axis.bar(x, frame['InitialSOC_MWh'], width, label='InitialSOC_MWh')
    axis.bar(x + width, frame['MinSOC_MWh'], width, label='MinSOC_MWh')
    axis.set_xticks(x)
    axis.set_xticklabels(frame['Region'])
    axis.set_ylabel('MWh')
    axis.set_title('各区域储能容量与SOC边界')
    axis.legend()
    save_figure(fig, 'eda_storage_capacity_soc.png')

    plot_correlation(
        frame,
        [
            'StorageCapacity_MWh', 'MinSOC_MWh', 'InitialSOC_MWh',
            'MaxChargePower_MW', 'MaxDischargePower_MW',
            'ChargeEfficiency', 'DischargeEfficiency',
            'MaxGridImport_MW', 'MaxGridExport_MW'
        ],
        'eda_storage_correlation.png',
        '储能参数相关性'
    )

    print_iqr_outliers(
        frame,
        [
            'StorageCapacity_MWh', 'InitialSOC_MWh',
            'MaxChargePower_MW', 'MaxDischargePower_MW',
            'MaxGridImport_MW', 'MaxGridExport_MW'
        ],
        label='储能参数'
    )


def analyze_workload(frame, power_mapping):
    print_section('任务数据有效性检查')
    before_count = len(frame)
    valid_mask = (
        frame['TaskID'].notna()
        & frame['TaskType'].isin(power_mapping['TaskType'])
        & frame['ArrivalHour'].between(0, 2399)
        & (frame['GPU_Demand'] > 0)
        & (frame['EstimatedDuration_min'] > 0)
        & (frame['EarliestStartHour'] >= frame['ArrivalHour'])
        & (frame['LatestFinishHour'] <= 2406)
        & (frame['LatestFinishHour'] > frame['EarliestStartHour'])
    )
    valid = frame.loc[valid_mask].copy()
    print(f'有效性筛选前行数: {before_count}')
    print(f'有效性筛选后行数: {len(valid)}')
    print(f'被筛除行数: {before_count - len(valid)}')

    print('各项有效性规则不满足数量:')
    rule_failures = {
        'TaskID缺失': frame['TaskID'].isna(),
        '任务类型不在功率映射中': ~frame['TaskType'].isin(power_mapping['TaskType']),
        'ArrivalHour不在0至2399': ~frame['ArrivalHour'].between(0, 2399),
        'GPU_Demand非正': frame['GPU_Demand'] <= 0,
        'EstimatedDuration_min非正': frame['EstimatedDuration_min'] <= 0,
        'EarliestStartHour早于ArrivalHour':
            frame['EarliestStartHour'] < frame['ArrivalHour'],
        'LatestFinishHour超过2406': frame['LatestFinishHour'] > 2406,
        '完成时限不晚于最早开工':
            frame['LatestFinishHour'] <= frame['EarliestStartHour'],
    }
    for rule, failed in rule_failures.items():
        print(f'{rule}: {int(failed.sum())}')

    valid['Duration_hour'] = valid['EstimatedDuration_min'] / 60.0
    valid['GPU_hour'] = valid['GPU_Demand'] * valid['Duration_hour']
    valid['SchedulingSlack_hour'] = (
        valid['LatestFinishHour']
        - valid['EarliestStartHour']
        - valid['Duration_hour']
    )
    valid = valid.merge(
        power_mapping[
            ['TaskType', 'GPU_Power_MW_per_EquivalentGPU']
        ],
        on='TaskType',
        how='left',
        validate='many_to_one'
    )
    valid['EstimatedEnergy_MWh'] = (
        valid['GPU_hour']
        * valid['GPU_Power_MW_per_EquivalentGPU']
    )

    print_section('任务总体描述统计')
    analysis_columns = [
        'GPU_Demand', 'EstimatedDuration_min', 'Duration_hour',
        'GPU_hour', 'MaxLatency_ms', 'SchedulingSlack_hour',
        'EstimatedEnergy_MWh'
    ]
    print(
        valid[analysis_columns].describe(
            percentiles=[0.01, 0.05, 0.25, 0.5, 0.75, 0.95, 0.99]
        ).T.to_string()
    )

    print('任务类型分布:')
    print(valid['TaskType'].value_counts(dropna=False).to_string())
    print('来源区域分布:')
    print(valid['SourceRegion'].value_counts(dropna=False).to_string())
    print('时延敏感度分布:')
    print(valid['DelaySensitivity'].value_counts(dropna=False).to_string())
    print('执行模式分布:')
    print(valid['ExecutionMode'].value_counts(dropna=False).to_string())

    grouped = valid.groupby(['SourceRegion', 'TaskType']).agg(
        TaskCount=('TaskID', 'count'),
        GPU_Demand_Total=('GPU_Demand', 'sum'),
        GPU_Demand_Mean=('GPU_Demand', 'mean'),
        GPU_Demand_Median=('GPU_Demand', 'median'),
        GPU_Demand_P95=('GPU_Demand', lambda x: x.quantile(0.95)),
        GPU_hour_Total=('GPU_hour', 'sum'),
        GPU_hour_Mean=('GPU_hour', 'mean'),
        Duration_min_Mean=('EstimatedDuration_min', 'mean'),
        Duration_min_Median=('EstimatedDuration_min', 'median'),
        Duration_min_P95=(
            'EstimatedDuration_min', lambda x: x.quantile(0.95)
        ),
        EstimatedEnergy_MWh_Total=('EstimatedEnergy_MWh', 'sum')
    )
    print_section('区域与任务类型组合统计')
    print(grouped.to_string())

    hourly = valid.groupby(
        ['ArrivalHour', 'SourceRegion', 'TaskType'], as_index=False
    ).agg(
        TaskCount=('TaskID', 'count'),
        Arrival_GPU_Demand=('GPU_Demand', 'sum'),
        Arrival_GPU_hour=('GPU_hour', 'sum'),
        EstimatedEnergy_MWh=('EstimatedEnergy_MWh', 'sum')
    )

    complete_index = pd.MultiIndex.from_product(
        [
            range(2400),
            sorted(valid['SourceRegion'].unique()),
            sorted(valid['TaskType'].unique())
        ],
        names=['ArrivalHour', 'SourceRegion', 'TaskType']
    )
    hourly_complete = hourly.set_index(
        ['ArrivalHour', 'SourceRegion', 'TaskType']
    ).reindex(complete_index, fill_value=0).reset_index()

    print_section('逐小时到达负荷统计')
    print(
        hourly_complete[
            ['TaskCount', 'Arrival_GPU_Demand', 'Arrival_GPU_hour']
        ].describe(
            percentiles=[0.01, 0.05, 0.25, 0.5, 0.75, 0.95, 0.99]
        ).T.to_string()
    )

    print('各区域任务到达小时覆盖范围:')
    coverage = valid.groupby('SourceRegion')['ArrivalHour'].agg(
        ['min', 'max', 'nunique']
    )
    print(coverage.to_string())

    by_hour_of_day = valid.assign(
        HourOfDay=valid['ArrivalHour'] % 24
    ).groupby(['HourOfDay', 'TaskType']).agg(
        TaskCount=('TaskID', 'count'),
        GPU_Demand_Total=('GPU_Demand', 'sum'),
        GPU_hour_Total=('GPU_hour', 'sum')
    )
    print_section('小时周期统计')
    print(by_hour_of_day.to_string())

    by_day = valid.assign(
        Day=valid['ArrivalHour'] // 24
    ).groupby(['Day', 'TaskType']).agg(
        TaskCount=('TaskID', 'count'),
        GPU_Demand_Total=('GPU_Demand', 'sum'),
        GPU_hour_Total=('GPU_hour', 'sum')
    )
    print_section('每日任务统计')
    print(by_day.to_string())

    print_section('逐小时GPU需求自相关')
    regional_hourly = hourly_complete.groupby(
        ['ArrivalHour', 'SourceRegion']
    )['Arrival_GPU_Demand'].sum().unstack(fill_value=0)
    for region in regional_hourly.columns:
        for lag in [1, 2, 3, 6, 12, 24, 48, 168]:
            value = regional_hourly[region].autocorr(lag=lag)
            print(f'Region={region}, lag={lag}, autocorrelation={value}')

    print_section('区域间逐小时到达GPU需求相关性')
    print(regional_hourly.corr().to_string())

    print_section('末端时域任务统计')
    final_tasks = valid[valid['ArrivalHour'].between(2376, 2399)]
    print(f'第2376至2399小时到达任务数: {len(final_tasks)}')
    print(
        final_tasks.groupby(['ArrivalHour', 'TaskType']).agg(
            TaskCount=('TaskID', 'count'),
            GPU_Demand_Total=('GPU_Demand', 'sum'),
            GPU_hour_Total=('GPU_hour', 'sum'),
            EstimatedEnergy_MWh_Total=('EstimatedEnergy_MWh', 'sum')
        ).to_string()
    )
    print(
        f'第2376至2399小时到达且LatestFinishHour大于2399的任务数: '
        f'{int((final_tasks["LatestFinishHour"] > 2399).sum())}'
    )
    print(
        f'全部任务中LatestFinishHour等于2406的任务数: '
        f'{int((valid["LatestFinishHour"] == 2406).sum())}'
    )
    print(
        f'调度松弛时间为负的任务数: '
        f'{int((valid["SchedulingSlack_hour"] < 0).sum())}'
    )

    plot_numeric_histograms(
        valid,
        [
            'GPU_Demand', 'EstimatedDuration_min', 'GPU_hour',
            'MaxLatency_ms', 'SchedulingSlack_hour', 'EstimatedEnergy_MWh'
        ],
        'eda_workload_numeric_distributions.png',
        '任务数值变量分布'
    )

    fig, axes = plt.subplots(1, 2, figsize=(15, 5))
    valid['TaskType'].value_counts().plot.bar(ax=axes[0], color='#4472C4')
    axes[0].set_title('任务类型数量分布')
    axes[0].set_xlabel('任务类型')
    axes[0].set_ylabel('任务数')
    axes[0].tick_params(axis='x', rotation=15)

    valid['SourceRegion'].value_counts().sort_index().plot.bar(
        ax=axes[1], color='#70AD47'
    )
    axes[1].set_title('来源区域任务数量分布')
    axes[1].set_xlabel('来源区域')
    axes[1].set_ylabel('任务数')
    save_figure(fig, 'eda_workload_category_distributions.png')

    fig, axis = plt.subplots(figsize=(16, 7))
    total_hourly = hourly_complete.groupby(
        ['ArrivalHour', 'TaskType']
    )['Arrival_GPU_Demand'].sum().unstack(fill_value=0)
    for task_type in total_hourly.columns:
        axis.plot(
            total_hourly.index,
            total_hourly[task_type],
            linewidth=0.8,
            label=task_type
        )
    axis.set_xlabel('到达小时')
    axis.set_ylabel('到达GPU需求')
    axis.set_title('各任务类型逐小时到达GPU需求')
    axis.legend()
    save_figure(fig, 'eda_hourly_gpu_demand_by_task_type.png')

    fig, axis = plt.subplots(figsize=(12, 6))
    hour_of_day_plot = by_hour_of_day['GPU_Demand_Total'].unstack(fill_value=0)
    hour_of_day_plot.plot(ax=axis, marker='o')
    axis.set_xlabel('日内小时')
    axis.set_ylabel('GPU需求总量')
    axis.set_title('任务到达GPU需求的日内周期')
    axis.set_xticks(range(24))
    save_figure(fig, 'eda_workload_hour_of_day_pattern.png')

    plot_correlation(
        valid,
        [
            'GPU_Demand', 'EstimatedDuration_min', 'Duration_hour',
            'GPU_hour', 'MaxLatency_ms', 'SchedulingSlack_hour',
            'EstimatedEnergy_MWh'
        ],
        'eda_workload_correlation.png',
        '任务数值变量相关性'
    )

    print_iqr_outliers(
        valid,
        [
            'GPU_Demand', 'EstimatedDuration_min', 'GPU_hour',
            'MaxLatency_ms', 'SchedulingSlack_hour',
            'EstimatedEnergy_MWh'
        ],
        group_column='TaskType',
        label='按任务类型的任务属性'
    )

    print_section('逐小时任务序列变化异常检测')
    print(
        '任务序列具有时序波动,异常检测使用逐小时一阶差分,'
        '不对原始逐小时负荷直接应用全局IQR.'
    )
    for region in regional_hourly.columns:
        differences = regional_hourly[region].diff().dropna()
        q1 = differences.quantile(0.25)
        q3 = differences.quantile(0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        anomaly_count = int(
            ((differences < lower) | (differences > upper)).sum()
        )
        print(
            f'Region={region}, 一阶差分Q1={q1}, 一阶差分Q3={q3}, '
            f'IQR={iqr}, 下界={lower}, 上界={upper}, '
            f'异常变化数量={anomaly_count}'
        )

    return valid, hourly_complete


def analyze_region_time_data(frame, gpu_info, storage_info):
    print_section('区域逐时数据完整性与边界检查')
    expected_hours = set(range(2407))
    expected_regions = set(gpu_info['Region'])
    actual_hours = set(frame['Hour'].dropna().astype(int))
    actual_regions = set(frame['Region'].dropna())

    print(f'实际Hour最小值: {frame["Hour"].min()}')
    print(f'实际Hour最大值: {frame["Hour"].max()}')
    print(f'Hour唯一值数量: {frame["Hour"].nunique()}')
    print(f'Region唯一值数量: {frame["Region"].nunique()}')
    print(f'缺失小时数量: {len(expected_hours - actual_hours)}')
    print(f'缺失区域数量: {len(expected_regions - actual_regions)}')
    print(f'重复Region-Hour组合数量: {int(frame.duplicated(["Region", "Hour"]).sum())}')

    counts = frame.groupby('Region')['Hour'].agg(['count', 'min', 'max', 'nunique'])
    print('各区域时点覆盖:')
    print(counts.to_string())

    period_counts = frame.groupby(['DataPeriod', 'Region']).size()
    print('DataPeriod与区域行数:')
    print(period_counts.to_string())

    print('第2400至2406小时分区记录数:')
    print(
        frame[frame['Hour'].between(2400, 2406)]
        .groupby(['Hour', 'Region'])
        .size()
        .to_string()
    )

    numeric_columns = frame.select_dtypes(include=np.number).columns.tolist()
    non_time_numeric = [column for column in numeric_columns if column != 'Hour']

    print_section('区域逐时数值变量分区描述统计')
    for column in non_time_numeric:
        summary = frame.groupby('Region')[column].agg(
            ['count', 'min', 'max', 'mean', 'median', 'std']
        )
        print(f'变量: {column}')
        print(summary.to_string())

    print_section('区域逐时类别变量分布')
    for column in ['PricePeriod', 'DemandResponseLevel', 'DataPeriod']:
        print(f'变量: {column}')
        print(frame.groupby(['Region', column]).size().to_string())

    merged = frame.merge(
        gpu_info[
            [
                'Region', 'Available_GPU', 'Max_IT_Power_MW',
                'Max_Facility_Power_MW', 'PUE'
            ]
        ],
        on='Region',
        how='left',
        validate='many_to_one'
    ).merge(
        storage_info[
            [
                'Region', 'StorageCapacity_MWh', 'MinSOC_MWh',
                'MaxChargePower_MW', 'MaxDischargePower_MW',
                'MaxGridImport_MW', 'MaxGridExport_MW'
            ]
        ],
        on='Region',
        how='left',
        validate='many_to_one'
    )

    merged['Calculated_Total_Load_MW'] = merged['IT_Load_MW'] * merged['PUE']
    merged['Total_Load_Difference_MW'] = (
        merged['Total_Load_MW'] - merged['Calculated_Total_Load_MW']
    )
    merged['Calculated_NetGridImport_MW'] = (
        merged['GridPurchase_MW'] - merged['GridSell_MW']
    )
    merged['NetGridImport_Difference_MW'] = (
        merged['NetGridImport_MW']
        - merged['Calculated_NetGridImport_MW']
    )
    merged['Renewable_Balance_Difference_MW'] = (
        merged['AvailableRenewable_MW']
        - merged['UsedRenewable_MW']
        - merged['RenewableCharge_MW']
        - merged['Curtailment_MW']
    )
    merged['Calculated_CarbonEmission_tCO2'] = (
        merged['GridPurchase_MW']
        * merged['CarbonIntensity_tCO2_per_MWh']
    )
    merged['CarbonEmission_Difference_tCO2'] = (
        merged['CarbonEmission_tCO2']
        - merged['Calculated_CarbonEmission_tCO2']
    )

    print_section('逐时计算口径复核')
    for column in [
        'Total_Load_Difference_MW',
        'NetGridImport_Difference_MW',
        'Renewable_Balance_Difference_MW',
        'CarbonEmission_Difference_tCO2'
    ]:
        print(f'变量: {column}')
        print(merged[column].describe().to_string())
        print(f'绝对差大于0.001的行数: {int((merged[column].abs() > 0.001).sum())}')

    boundary_checks = {
        'GPU利用率低于0':
            merged['GPU_Utilization_Percent'] < 0,
        'GPU利用率超过100':
            merged['GPU_Utilization_Percent'] > 100,
        'IT负荷超过区域IT功率上限':
            merged['IT_Load_MW'] > merged['Max_IT_Power_MW'],
        '设施负荷超过区域设施功率上限':
            merged['Total_Load_MW'] > merged['Max_Facility_Power_MW'],
        'SOC低于最小值':
            merged['SOC_MWh'] < merged['MinSOC_MWh'],
        'SOC超过储能容量':
            merged['SOC_MWh'] > merged['StorageCapacity_MWh'],
        '充电功率超过上限':
            merged['ChargePower_MW'] > merged['MaxChargePower_MW'],
        '放电功率超过上限':
            merged['DischargePower_MW'] > merged['MaxDischargePower_MW'],
        '电网购电超过上限':
            merged['GridPurchase_MW'] > merged['MaxGridImport_MW'],
        '电网售电超过上限':
            merged['GridSell_MW'] > merged['MaxGridExport_MW'],
        '可用新能源为负':
            merged['AvailableRenewable_MW'] < 0,
        '弃电为负':
            merged['Curtailment_MW'] < 0,
        '充电和放电同时为正':
            (merged['ChargePower_MW'] > 0)
            & (merged['DischargePower_MW'] > 0),
        '购电和售电同时为正':
            (merged['GridPurchase_MW'] > 0)
            & (merged['GridSell_MW'] > 0),
    }
    print_section('逐时运行边界违规统计')
    for rule, mask in boundary_checks.items():
        print(f'{rule}: {int(mask.fillna(False).sum())}')

    print_section('区域累计能源与排放指标')
    regional_totals = merged.groupby('Region').agg(
        AvailableRenewable_MWh=('AvailableRenewable_MW', 'sum'),
        UsedRenewable_MWh=('UsedRenewable_MW', 'sum'),
        RenewableCharge_MWh=('RenewableCharge_MW', 'sum'),
        Curtailment_MWh=('Curtailment_MW', 'sum'),
        GridPurchase_MWh=('GridPurchase_MW', 'sum'),
        GridSell_MWh=('GridSell_MW', 'sum'),
        NetGridImport_MWh=('NetGridImport_MW', 'sum'),
        CarbonEmission_tCO2=('CarbonEmission_tCO2', 'sum'),
        TotalLoad_MWh=('Total_Load_MW', 'sum'),
        ITLoad_MWh=('IT_Load_MW', 'sum')
    )
    regional_totals['RenewableUtilizationRatio'] = (
        regional_totals['UsedRenewable_MWh']
        + regional_totals['RenewableCharge_MWh']
        + regional_totals['GridSell_MWh']
    ) / regional_totals['AvailableRenewable_MWh']
    print(regional_totals.to_string())

    print_section('系统逐小时峰值与波动统计')
    system_hourly = merged.groupby('Hour').agg(
        Total_Load_MW=('Total_Load_MW', 'sum'),
        IT_Load_MW=('IT_Load_MW', 'sum'),
        GridPurchase_MW=('GridPurchase_MW', 'sum'),
        GridSell_MW=('GridSell_MW', 'sum'),
        NetGridImport_MW=('NetGridImport_MW', 'sum'),
        AvailableRenewable_MW=('AvailableRenewable_MW', 'sum'),
        UsedRenewable_MW=('UsedRenewable_MW', 'sum'),
        CarbonEmission_tCO2=('CarbonEmission_tCO2', 'sum')
    )
    print(system_hourly.describe().T.to_string())
    for column in system_hourly.columns:
        peak_hour = system_hourly[column].idxmax()
        trough_hour = system_hourly[column].idxmin()
        print(
            f'变量={column}, 最大值={system_hourly.loc[peak_hour, column]}, '
            f'最大值小时={peak_hour}, 最小值={system_hourly.loc[trough_hour, column]}, '
            f'最小值小时={trough_hour}'
        )

    print_section('区域逐时序列变化异常检测')
    print(
        '区域逐时数据具有明显时序趋势和周期,异常检测使用按区域排序后的一阶差分,'
        '不对原始温度或状态类数值应用全局IQR.'
    )
    diff_columns = [
        'ElectricityPrice_CNY_per_MWh',
        'CarbonIntensity_tCO2_per_MWh',
        'GPU_Utilization_Percent',
        'AvailableRenewable_MW',
        'IT_Load_MW',
        'Total_Load_MW',
        'NetGridImport_MW',
        'SOC_MWh'
    ]
    sorted_frame = merged.sort_values(['Region', 'Hour']).copy()
    for column in diff_columns:
        differences = sorted_frame.groupby('Region')[column].diff()
        for region in sorted_frame['Region'].drop_duplicates():
            region_mask = sorted_frame['Region'] == region
            values = differences.loc[region_mask].dropna()
            q1 = values.quantile(0.25)
            q3 = values.quantile(0.75)
            iqr = q3 - q1
            lower = q1 - 1.5 * iqr
            upper = q3 + 1.5 * iqr
            anomaly_count = int(((values < lower) | (values > upper)).sum())
            print(
                f'变量={column}, Region={region}, 差分Q1={q1}, '
                f'差分Q3={q3}, IQR={iqr}, 下界={lower}, 上界={upper}, '
                f'异常变化数量={anomaly_count}'
            )

    plot_numeric_histograms(
        merged,
        [
            'ElectricityPrice_CNY_per_MWh',
            'CarbonIntensity_tCO2_per_MWh',
            'GPU_Utilization_Percent',
            'AvailableRenewable_MW',
            'IT_Load_MW',
            'Total_Load_MW',
            'NetGridImport_MW',
            'CarbonEmission_tCO2',
            'SOC_MWh'
        ],
        'eda_region_time_numeric_distributions.png',
        '区域逐时主要数值变量分布'
    )

    plot_correlation(
        merged,
        [
            'ElectricityPrice_CNY_per_MWh',
            'SellPrice_CNY_per_MWh',
            'CarbonIntensity_tCO2_per_MWh',
            'GPU_Utilization_Percent',
            'AvailableRenewable_MW',
            'UsedRenewable_MW',
            'Curtailment_MW',
            'IT_Load_MW',
            'Total_Load_MW',
            'GridPurchase_MW',
            'GridSell_MW',
            'NetGridImport_MW',
            'CarbonEmission_tCO2',
            'SOC_MWh'
        ],
        'eda_region_time_correlation.png',
        '区域逐时变量相关性'
    )

    for column, ylabel, filename in [
        ('Total_Load_MW', 'MW', 'eda_total_load_timeseries.png'),
        ('NetGridImport_MW', 'MW', 'eda_net_grid_import_timeseries.png'),
        ('SOC_MWh', 'MWh', 'eda_soc_timeseries.png'),
        (
            'AvailableRenewable_MW',
            'MW',
            'eda_available_renewable_timeseries.png'
        )
    ]:
        pivot = merged.pivot(index='Hour', columns='Region', values=column)
        fig, axis = plt.subplots(figsize=(16, 7))
        for region in pivot.columns:
            axis.plot(pivot.index, pivot[region], linewidth=0.7, label=region)
        axis.set_xlabel('小时')
        axis.set_ylabel(ylabel)
        axis.set_title(f'{column}逐时变化')
        axis.legend(ncol=3)
        save_figure(fig, filename)

    fig, axis = plt.subplots(figsize=(16, 7))
    axis.plot(
        system_hourly.index,
        system_hourly['NetGridImport_MW'],
        label='系统净购电',
        linewidth=0.9
    )
    axis.plot(
        system_hourly.index,
        system_hourly['Total_Load_MW'],
        label='系统设施负荷',
        linewidth=0.9
    )
    axis.set_xlabel('小时')
    axis.set_ylabel('MW')
    axis.set_title('系统设施负荷与净购电')
    axis.legend()
    save_figure(fig, 'eda_system_load_and_net_import.png')

    return merged, system_hourly


def analyze_cross_dataset(
    workload, hourly_workload, region_time, gpu_info, latency
):
    print_section('跨数据集一致性检查')
    workload_regions = set(workload['SourceRegion'])
    gpu_regions = set(gpu_info['Region'])
    time_regions = set(region_time['Region'])
    latency_from = set(latency['FromRegion'])
    latency_to = set(latency['ToRegion'])

    print(f'任务来源区域集合: {sorted(workload_regions)}')
    print(f'GPU区域集合: {sorted(gpu_regions)}')
    print(f'逐时数据区域集合: {sorted(time_regions)}')
    print(f'时延来源区域集合: {sorted(latency_from)}')
    print(f'时延目标区域集合: {sorted(latency_to)}')
    print(f'任务来源区域不在GPU信息中的数量: {len(workload_regions - gpu_regions)}')
    print(f'GPU区域不在逐时数据中的数量: {len(gpu_regions - time_regions)}')

    latency_pairs = set(zip(latency['FromRegion'], latency['ToRegion']))
    expected_pairs = {
        (source, target)
        for source in gpu_regions
        for target in gpu_regions
    }
    print(f'网络时延缺失有向区域对数量: {len(expected_pairs - latency_pairs)}')

    baseline = region_time[
        region_time['Hour'].between(0, 2399)
    ].pivot(
        index='Hour',
        columns='Region',
        values='Baseline_AI_IT_Load_MW'
    )
    arrival_gpu = hourly_workload.groupby(
        ['ArrivalHour', 'SourceRegion']
    )['Arrival_GPU_Demand'].sum().unstack(fill_value=0)

    common_hours = baseline.index.intersection(arrival_gpu.index)
    common_regions = baseline.columns.intersection(arrival_gpu.columns)
    cross_correlation = pd.DataFrame(index=common_regions, columns=common_regions)
    for source_region in common_regions:
        for load_region in common_regions:
            cross_correlation.loc[source_region, load_region] = (
                arrival_gpu.loc[common_hours, source_region].corr(
                    baseline.loc[common_hours, load_region]
                )
            )

    print('来源区域逐小时到达GPU需求与区域基线AI负荷相关性:')
    print(cross_correlation.to_string())


def main():
    print_section('EDA开始')
    print(f'数据目录: {DATA_DIR}')
    print(f'图表输出目录: {FIGURE_DIR}')

    workbooks = {}
    for name, path in FILES.items():
        workbooks[name] = load_excel_workbook(name, path)

    gpu_info = workbooks['GPU_information']['GPU中心基础情况']
    network_latency = workbooks['network_latency']['network_latency']
    power_mapping = workbooks['power_mapping']['任务功率映射']
    region_time = workbooks['region_time_data']['region_time_data']
    storage_info = workbooks['storage_information']['storage_information']
    workload = workbooks['workload_trace']['Sheet1']

    analyze_gpu_information(gpu_info)
    analyze_network_latency(network_latency)
    analyze_power_mapping(power_mapping)
    analyze_storage_information(storage_info)

    valid_workload, hourly_workload = analyze_workload(
        workload, power_mapping
    )
    analyzed_region_time, system_hourly = analyze_region_time_data(
        region_time, gpu_info, storage_info
    )
    analyze_cross_dataset(
        valid_workload,
        hourly_workload,
        analyzed_region_time,
        gpu_info,
        network_latency
    )

    print_section('生成图表清单')
    generated_figures = sorted(FIGURE_DIR.glob('eda_*.png'))
    print(f'图表数量: {len(generated_figures)}')
    for figure in generated_figures:
        print(str(figure))

    print_section('EDA完成')


if __name__ == '__main__':
    main()